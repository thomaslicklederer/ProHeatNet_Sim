import pickle
import math
import pprint
import openpyxl as pyxl

def excelcompare(case_name, workbook_path, iteration):
    '''code by M.Sc. Thomas Licklederer, Technical University of Munich, MSE, all
    rights reserved'''

    # load data from pickle file
    myfile = open('results/results.pkl', 'rb')
    myresults = pickle.load(myfile)
    myfile.close()
    PSM = myresults['PSM']
    v_vec = myresults['v_vec']
    e_vec = myresults['e_vec']
    topology = myresults['topology']
    scenario = myresults['scenario']
    solutions = myresults['solutions']

    # load excel file
    workbook = pyxl.load_workbook(workbook_path)
    sheet = workbook.active
    
    # write case name
    casecell = sheet.cell(row=iteration, column=1)
    casecell.value = case_name

    # content
    myrow = 0
    mycol = 2

    # transferred heat
    for P in PSM:
        headercell = sheet.cell(row=1, column=mycol)   
        entrycell = sheet.cell(row=iteration, column=mycol)  
        
        header = 'dotQ_%1.0f' % P
        entry = -solutions['Q_trnsf2'][P]
        
        headercell.value = header
        entrycell.value = float(entry)
        
        mycol +=1
        
    # sum of heat losses
    sum_Q_loss = 0
    for key in solutions['Q_loss'].keys():
        sum_Q_loss += solutions['Q_loss'][key]
    headercell = sheet.cell(row=1, column=mycol)   
    entrycell = sheet.cell(row=iteration, column=mycol)  
    header = 'sum_dotQ_loss'
    entry = sum_Q_loss
    headercell.value = header
    entrycell.value = float(entry)
    mycol +=1

    # network (primary side) temperatures at prosumers
    for P in PSM:
        headercell = sheet.cell(row=1, column=mycol)   
        entrycell = sheet.cell(row=iteration, column=mycol)  
        header = 'T_%1.0f_prim_hot' % P
        key= ('%1.0fh' % P, '%1.0fc' % P)
        entry = solutions['T'][key]
        headercell.value = header
        entrycell.value = float(entry)
        mycol += 1
        
        headercell = sheet.cell(row=1, column=mycol)   
        entrycell = sheet.cell(row=iteration, column=mycol)  
        header = 'T_%1.0f_prim_cold' % P
        key= ('%1.0fc' % P, '%1.0fh' % P)
        entry = solutions['T'][key]
        headercell.value = header
        entrycell.value = float(entry)
        mycol += 1

    # prosumer (secondary side) temperatures at prosumers
    for P in PSM:
        headercell = sheet.cell(row=1, column=mycol)   
        entrycell = sheet.cell(row=iteration, column=mycol)  
        header = 'T_%1.0f_sec_hot' % P
        key= 'PSM%1.0fh' % P
        entry = solutions['T'][key]
        headercell.value = header
        entrycell.value = float(entry)
        mycol += 1
        
        headercell = sheet.cell(row=1, column=mycol)   
        entrycell = sheet.cell(row=iteration, column=mycol)  
        header = 'T_%1.0f_sec_cold' % P
        key= 'PSM%1.0fc' % P
        entry = solutions['T'][key]
        headercell.value = header
        entrycell.value = float(entry)
        mycol += 1

    # volume flows through prosumer substations at primary side
    for P in PSM:
        headercell = sheet.cell(row=1, column=mycol)   
        entrycell = sheet.cell(row=iteration, column=mycol)  
        header = 'dotV_%1.0f_prim' % P
        key= ('%1.0fh' % P, '%1.0fc' % P)
        entry = solutions['dotV'][key]
        headercell.value = header
        entrycell.value = float(entry)
        mycol += 1
     
    # volume flows through prosumer substations at secondary side
    for P in PSM:
        headercell = sheet.cell(row=1, column=mycol)   
        entrycell = sheet.cell(row=iteration, column=mycol)  
        header = 'dotV_%1.0f_sec' % P
        entry = solutions['dotV'][P]
        headercell.value = header
        entrycell.value = float(entry)
        mycol += 1

    # pressure differences over substations primary side
    for P in PSM:
        headercell = sheet.cell(row=1, column=mycol)   
        entrycell = sheet.cell(row=iteration, column=mycol)  
        header = 'Deltap_%1.0f_prim' % P
        key= ('%1.0fh' % P, '%1.0fc' % P)
        entry = solutions['Deltap'][key]
        headercell.value = header
        entrycell.value = float(entry)
        mycol += 1

    #############
    mycol += 1
    #############

    # volume flows in network
    for key in solutions['dotV'].keys():
        headercell = sheet.cell(row=1, column=mycol)   
        entrycell = sheet.cell(row=iteration, column=mycol)  
        try:
            header = 'dotV_%s_%s' % (key[0], key[1])
        except:
            header = 'dotV_%s' % key
        entry = solutions['dotV'][key]
        headercell.value = header
        entrycell.value = float(entry)
        mycol += 1
        
    # pressure differences in network
    for key in solutions['Deltap'].keys():
        headercell = sheet.cell(row=1, column=mycol)   
        entrycell = sheet.cell(row=iteration, column=mycol)  
        if len(key)>1:
            header = 'Deltap_%s_%s' % (key[0], key[1])
        else:
            header = 'Deltap_%s' % key
        entry = solutions['Deltap'][key]
        headercell.value = header
        entrycell.value = float(entry)
        mycol += 1
        
    # temperatures in network
    for key in solutions['T'].keys():
        headercell = sheet.cell(row=1, column=mycol)   
        entrycell = sheet.cell(row=iteration, column=mycol)  
        if not 'PSM' in key:
            header = 'T_%s_%s' % (key[0], key[1])
        else:
            header = 'T_%s' % key
        entry = solutions['T'][key]
        headercell.value = header
        entrycell.value = float(entry)
        mycol += 1
        
    workbook.save(workbook_path)
